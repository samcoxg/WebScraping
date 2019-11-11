from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import pyautogui, sys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
import os
import time
import xlrd
import xlwt



#browser = webdriver.Safari()
def webpage_initiate():
    global browser
    browser = webdriver.Chrome(ChromeDriverManager().install())
    browser.get("https://link.springer.com/journal/volumesAndIssues/41303")
    browser.maximize_window()
    

issueCount = 5
volumeCount = 26
year = 2019
writeCount = 2

def link_destroy(): 
    articleSelection = 1
    global iCount, volumeCount, issueCount, writeCount, year

    wb=Workbook()
    filepath="abstract_tests.xlsx"
    wb=load_workbook(filepath)
    sheet=wb.active
    sheet['A1'] = 1
    sheet.cell(row=1, column=1).value = "Title"
    sheet.cell(row=1, column=2).value = "Abstract"
    sheet.cell(row=1, column=3).value = "Authors"
    sheet.cell(row=1, column=4).value = "Keywords"
    sheet.cell(row=1, column=5).value = "Volume"
    sheet.cell(row=1, column=6).value = "Issue"
    sheet.cell(row=1, column=7).value = "Year"
    
    

    while(iCount < 1000):
        
        if(issueCount == 0):
            issueCount = 6
            volumeCount = volumeCount - 1
            year = year - 1
        if(volumeCount == 0):
            iCount = 0
        articleReach = True
        
        #click to expand the volume
        #There may be a more direct way to go to the page via links
        browser.get("https://link.springer.com/journal/volumesAndIssues/41303")
        
        #run while still able to reach the articles
        while(articleReach == True):
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            #get this url based on the current issue and volume being searched for
            browser.get("https://link.springer.com/journal/41303/{}/{}".format(volumeCount, issueCount))       
            try:
                #wait for the presence of an element on the page 
                WebDriverWait(browser, 45).until(EC.presence_of_element_located((By.ID, "enumeration")))
                #then scroll down to render the entire page
                browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                try:
                    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    #select an article on the page
                    browser.find_element_by_xpath("//*[@id='kb-nav--main']/div[3]/ol/li[{}]/div/h3/a".format(articleSelection)).click()
                except:
                    #add to article selection and try again
                    articleSelection = articleSelection + 1
                    browser.find_element_by_xpath("//*[@id='kb-nav--main']/div[3]/ol/li[{}]/div/h3/a".format(articleSelection)).click()
            except:
                articleReach = False
                print("browser.find failed")
                articleSelection = 1
                break
            try:
                WebDriverWait(browser, 25).until(EC.presence_of_element_located((By.CLASS_NAME, "Heading")))
            except:
                articleReach = False
                print("abstract.find failed")
                articleSelection = 1
                break
            try:
                sheet.cell(row=writeCount, column=1).value = browser.find_element_by_xpath("//*[@id='main-content']/div/div/article/div/div[1]/div[2]/h1").text
                sheet.cell(row=writeCount, column=2).value = browser.find_element_by_xpath("//*[@id='Abs1']").text
                sheet.cell(row=writeCount, column=5).value = volumeCount
                sheet.cell(row=writeCount, column=6).value = issueCount
                try:
                    sheet.cell(row=writeCount, column=3).value = browser.find_element_by_xpath("//*[@id='authors']/ul").text
                    sheet.cell(row=writeCount, column=4).value = browser.find_element_by_xpath("//*[@id='main-content']/div/div/article/div/div[2]").text
                    sheet.cell(row=writeCount, column=7).value = browser.find_element_by_xpath("//*[@id='enumeration']/p[2]/span[1]/time").text
                except:
                    sheet.cell(row=writeCount, column=3).value = 'NA'
                    sheet.cell(row=writeCount, column=4).value = 'NA'
                    sheet.cell(row=writeCount, column=7).value = 'NA'
            except:
                articleReach = False
                print("abstract.find failed")
                articleSelection = 1
                break
            writeCount = writeCount + 1
            articleSelection = articleSelection + 1

        issueCount = issueCount - 1
        iCount = iCount + 1
        wb.save(filepath)

        global myMax 
        myMax = 10
        
        
myMax = 0       

while(myMax < 1):
    iCount = 0
    time.sleep(3)
    webpage_initiate()
    link_destroy()
    myMax = myMax + 1
    




