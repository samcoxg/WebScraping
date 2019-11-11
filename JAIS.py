from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import pyautogui, sys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
import os
import time
import xlrd
import xlwt



#browser = webdriver.Safari()
def webpage_initiate():
    global browser
    browser = webdriver.Chrome(ChromeDriverManager().install())
    browser.get("https://aisel.aisnet.org/jais/")
    browser.maximize_window()
    

issueCount = 12
volumeCount = 20
year = 2019
writeCount = 2

def link_destroy(): 
    articleSelection = 1
    global iCount, volumeCount, issueCount, writeCount, year

    wb=Workbook()
    filepath="abstract_tests.xlsx"
    wb=load_workbook(filepath)
    sheet=wb.active
    sheet['A1'] = 1
    sheet.cell(row=1, column=1).value = "Title"
    sheet.cell(row=1, column=2).value = "Abstract"
    sheet.cell(row=1, column=3).value = "Authors"
    sheet.cell(row=1, column=4).value = "Keywords"
    sheet.cell(row=1, column=5).value = "Volume"
    sheet.cell(row=1, column=6).value = "Issue"
    sheet.cell(row=1, column=7).value = "Year"
    sheet.cell(row=1, column=8).value = "University"
    
    

    while(iCount < 1000):
        
        if(issueCount == 0):
            issueCount = 12
            volumeCount = volumeCount - 1
            year = year - 1
        if(volumeCount == 0):
            iCount = 30

        articleReach = True
        browser.get("https://aisel.aisnet.org/jais/vol{}/iss{}".format(volumeCount,issueCount))
        
        while(articleReach == True):
            browser.get("https://aisel.aisnet.org/jais/vol{}/iss{}".format(volumeCount,issueCount))
            try:
                WebDriverWait(browser, 45).until(EC.presence_of_element_located((By.CLASS_NAME, "searchbutton")))
                browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                browser.find_element_by_xpath("//*[@id='main']/div[3]/div[{}]/p[2]/a".format(articleSelection)).click()
            except:
                articleReach = False
                print("browser.find failed")
                articleSelection = 1
                break
            try:
                WebDriverWait(browser, 25).until(EC.presence_of_element_located((By.ID, "abstract")))
            except:
                articleReach = False
                print("abstract.find failed")
                articleSelection = 1
                break
            sheet.cell(row=writeCount, column=1).value = browser.find_element_by_id("title").text
            sheet.cell(row=writeCount, column=2).value = browser.find_element_by_id("abstract").text
            sheet.cell(row=writeCount, column=3).value = browser.find_element_by_id("authors").text
            sheet.cell(row=writeCount, column=4).value = "NA"
            sheet.cell(row=writeCount, column=5).value = volumeCount
            sheet.cell(row=writeCount, column=6).value = issueCount
            sheet.cell(row=writeCount, column=7).value = year
            sheet.cell(row=2, column=8).value = "See author column"
            writeCount = writeCount + 1
            articleSelection = articleSelection + 1

        issueCount = issueCount - 1
        iCount = iCount + 1
        wb.save(filepath)

        global myMax 
        myMax = 10
        
        
myMax = 0       

while(myMax < 1):
    iCount = 0
    time.sleep(3)
    webpage_initiate()
    link_destroy()
    myMax = myMax + 1
    




