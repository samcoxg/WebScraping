#Import all standard packages for scraping, automation, and excel access and storage
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
import os, time, xlrd, xlwt, pyautogui, sys



#Initiate the webpage and specify the browser (Chrome webdriver shouldn't need to be installed, since it is included on line 19)
def webpage_initiate():
    global browser
    browser = webdriver.Chrome(ChromeDriverManager().install())
    browser.get("link to main page to begin loading")
    browser.maximize_window()
    
#Specify global variables to read through student data
studentCount = #someNumber
pageCount = #someNumber



def link_destroy(): 
    articleSelection = 1
    global iCount, studentCount, pageCount

    wb=Workbook()
    filepath="spreadsheet1.xlsx"
    wb=load_workbook(filepath)
    sheet=wb.active
    sheet['A1'] = 1
    sheet.cell(row=1, column=1).value = "field"
    sheet.cell(row=1, column=2).value = "field"
    sheet.cell(row=1, column=3).value = "field"
    sheet.cell(row=1, column=4).value = "field"
    sheet.cell(row=1, column=5).value = "field"
    sheet.cell(row=1, column=6).value = "field"
    sheet.cell(row=1, column=7).value = "field"
    
    
    #run while loop to iterate through the page. The page should be stable and the same accross all students, but exceptions are built in to prevent an entire program failure
    while(iCount < 1000):
        if(issueCount == 0):
            studentCount = #someNume
            pageCount = pageCount - 1
            year = year - 1
        if(volumeCount == 0):
            iCount = 0
        articleReach = True
        
        
        browser.get("main page url here")
        
        #run while still able to reach the students
        while(articleReach == True):
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")#scroll while render all data on the page in case of error
            #get this url based on the current student and page being searched for
            browser.get("url here".format(volumeCount, issueCount))    
            try:
                #wait for the presence of an element on the page 
                WebDriverWait(browser, 45).until(EC.presence_of_element_located((By.ID, "id")))
                #then scroll down to render the entire page again to prevent missed data
                browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            except:
                studentReach = False
                print("browser.find failed")
                studentSelection = 1
                break
            try:
                sheet.cell(row=writeCount, column=1).value = browser.find_element_by_xpath("xpath").text
                sheet.cell(row=writeCount, column=2).value = browser.find_element_by_xpath("xpath").text
                sheet.cell(row=writeCount, column=5).value = studentCount
                sheet.cell(row=writeCount, column=6).value = pageCount
                sheet.cell(row=writeCount, column=3).value = browser.find_element_by_xpath("xpath").text
                sheet.cell(row=writeCount, column=4).value = browser.find_element_by_xpath("xpath").text
                sheet.cell(row=writeCount, column=7).value = browser.find_element_by_xpath("xpath").text
            except:
                studentReach = False
                print("page.find failed")
                studentSelection = 1
                break
            writeCount = writeCount + 1
            studentSelection = studentSelection + 1

        studentSelection = studentSelection + 1
        iCount = iCount + 1
        #set the save
        wb.save(filepath)
        global myMax 
        myMax = 10
        
#set to zero and end program when while loop ends.        
myMax = 0       
while(myMax < 1):
    #pause, then close the program, 
    time.sleep(3)
    webpage_terminate()
    print("Messate about how page has either failed or ended.)
    




